import json
import openpyxl
from docx import Document

#open xls file and read values on first sheet
excel_doc = openpyxl.load_workbook('task_7_2_3.xlsx')
b = excel_doc.sheetnames
c = 'A1'
i = 1
sheet = excel_doc[b[0]]
while sheet[c].value != None:
    i += 1
    c = 'A' + str(i)
    print(c)
count_people = i - 1
sheet2 = excel_doc[b[1]]

#make second sheet with new structure of FIO
for i in range(1,(count_people+1)):
    num_b = 'B' + str(i)
    num_a = "A" + str(i)
    sheet2[num_b].value = sheet[num_a].value + " " + sheet[num_b].value
    col = 3
    while sheet.cell(row=i,column=col).value != 'end':
        sheet2.cell(row=i,column=col).value = sheet.cell(row=i,column=col).value
        col+=1
    letter_col = openpyxl.utils.get_column_letter((col-1))
    formula = '=СУММ(C'+str(i)+':'+letter_col+str(i)+')/'+str((col-3))
    sheet2.cell(row=i, column=1).value = formula
excel_doc.save('task_7_2_3.xlsx')

#count students
students = []
for i in range(1,(count_people+1)):
    num_b = 'B' + str(i)
    students.append(sheet2[num_b].value)

#count marks & average marks
def count_av(i,col=3):
    marks=[]
    a=0
    sum_marks=0
    while sheet2.cell(row=i,column=col).value != None:
        marks.append(sheet2.cell(row=i,column=col).value)
        col+=1
        a+=1
    for mark in marks:
        sum_marks+=float(mark)
    average_mark=float(sum_marks/a)
    return(marks,average_mark)

#import information in Json
marks=[]
average_marks=0
dict_for_json={}
dict_fio={}
for i in range(1,(count_people+1)):
    num_b = 'B' + str(i)
    dict_stud={}
    marks, average_mark = count_av(i)
    dict_stud['marks']=marks
    dict_stud['average_mark'] = average_mark
    average_marks+=average_mark
    dict_fio[sheet2[num_b].value] = dict_stud
dict_for_json['students'] = dict_fio
dict_for_json['total_aver_mark'] = float(average_marks/count_people)
print (dict_for_json['students'])
print (dict_for_json['total_aver_mark'])
with open ('students.json','w') as stud_json:
    json.dump(dict_for_json, stud_json, indent= 4)
print (json.dumps(dict_for_json, indent= 4))
